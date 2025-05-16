# TODO update paths


import json
import glob
import os
from openpyxl import load_workbook



# Directory containing Excel files
directory = "analysis/HV V2/"

judgments = {}

# Loop through each file in the directory
for filename in os.listdir(directory):
    if filename.endswith(".xlsx"):
        filepath = os.path.join(directory, filename)
        workbook = load_workbook(filename=filepath, data_only=True)
        sheetnames = workbook.sheetnames[4:8]  

        column_to_scan = 'B'

        video_id = filename[:-5]

        judgments[video_id] = {}

        print(f"\nFile: {filename}")
        for sheetname in sheetnames:
            sheet = workbook[sheetname]
            print(f"  Sheet: {sheetname}")

            row = 1
            end_count = 0
            in_reference = False
            reason = ""
            while True:
                if end_count > 4: break
                cell = f"{column_to_scan}{row}"
                value = sheet[cell].value
                if value == "Question":
                    in_reference = True
                    end_count = 0
                    reason = sheet['C' + str(row-2)].value
                    row += 1
                    judgments[video_id][reason] = {}
                if value == None:
                    in_reference = False
                    end_count += 1
                
                if in_reference:
                    reference_id = sheet['B' + str(row)].value
                    judgment = sheet['D' + str(row)].value

                    if judgment != None and int(judgment) >= 2:
                        judgments[video_id][reason][reference_id] = judgment                    
                row += 1

        new_lines = []
        with open("output/" + video_id + "/jir_references_relevance_score.jsonl", "r") as f:
            for line in f:
                json_line = json.loads(line)
                reason = json_line["reason"]

                old_references = json_line["references"].get("document_relevance_score_old", None)
                if not old_references:
                    old_references = json_line["references"]["document_relevance_score"]

                
                new_references = {}

                for entry in old_references:
                    if entry in judgments[video_id].get(reason, {}):
                        new_references[entry] = 3
                    elif old_references[entry] == 3:
                        new_references[entry] = 2
                    else:
                        new_references[entry] = 1

                json_line["references"]["document_relevance_score_old"] = old_references
                json_line["references"]["document_relevance_score"] = new_references
                new_lines.append(json_line)


        with open("output/" + video_id + "/jir_references_relevance_score.jsonl", "w") as f:
            for line in new_lines:
                f.write(json.dumps(line) + "\n")